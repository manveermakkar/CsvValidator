"""
core/reporter.py  —  Phase 2 (Refactored)
------------------------------------------
Generates three output formats:

1. Excel (.xlsx)  — openpyxl with red/green cell fills
                    Sheets:
                      1. <name_a>         — coloured data (File A)
                      2. <name_b>         — coloured data (File B)
                      3. Summary          — aggregate statistics
                      4. Mismatch Detail  — per-cell diff with primary key column
                      5. Fuzzy Match Log  — (only when fuzzy alignment is used)
2. HTML  (.html)  — self-contained inline-CSS coloured table
3. Text  (.txt)   — human-readable mismatch summary

Phase 2 additions
-----------------
* ``name_a`` / ``name_b`` parameters replace generic "File A" / "File B" labels
  throughout all three outputs.
* ``match_log`` (optional) — adds the Fuzzy Match Log sheet / TXT section.
* ``mismatch_list`` (optional) — adds the Mismatch Detail sheet with pk_label.
* ``pk_label`` — human-readable column header for the primary-key column
  e.g. "Primary Key (order_id)" or "Composite Key (region + product_id)".
"""

import io
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.aligner import STATUS_COL
from core.comparator import summary_stats


# ── Colour constants ─────────────────────────────────────────────────────────
_GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
_AMBER_FILL   = PatternFill("solid", fgColor="FFEB9C")
_BLUE_FILL    = PatternFill("solid", fgColor="DDEBF7")   # fuzzy match row
_TEAL_FILL    = PatternFill("solid", fgColor="D9EAD3")   # primary-key column
_GREY_FILL    = PatternFill("solid", fgColor="F2F2F2")   # unmatched row
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_PK_HDR_FILL  = PatternFill("solid", fgColor="274E13")   # PK column header (dark green)
_HEADER_FONT  = Font(color="FFFFFF", bold=True)
_FUZZY_FONT   = Font(color="375623", bold=True, italic=True)
_THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_ROWNR_FILL   = PatternFill("solid", fgColor="E2EFDA")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def to_excel(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    diff_mask: pd.DataFrame,
    stats: dict,
    name_a: str = "File A",
    name_b: str = "File B",
    match_log: list[dict] | None = None,
    mismatch_list: list[dict] | None = None,
    pk_label: str = "Primary Key",
) -> bytes:
    """
    Build an Excel workbook in memory and return raw bytes.

    Parameters
    ----------
    df_a, df_b     : Aligned DataFrames (already colour-mapped)
    diff_mask      : Boolean DataFrame — True where cells differ
    stats          : Summary statistics dict from summary_stats()
    name_a, name_b : Display names for each file
    match_log      : (optional) Fuzzy match log → adds Sheet 5
    mismatch_list  : (optional) List of mismatch dicts → adds Sheet 4
                     Each dict: {row, column, value_a, value_b,
                                 primary_key (optional)}
    pk_label       : Column header for the primary-key column in Sheet 4
                     e.g. "Primary Key (order_id)" or
                          "Composite Key (region + product_id)"

    Sheets:
      1. <name_a>        — coloured data from File A
      2. <name_b>        — coloured data from File B
      3. Summary         — match statistics + per-column counts
      4. Mismatch Detail — per-cell diff with primary-key column
      5. Fuzzy Match Log — (only when match_log is provided)
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: File A ───────────────────────────────────────────────────
    ws_a        = wb.active
    ws_a.title  = _excel_sheet_name(name_a)
    _write_coloured_sheet(ws_a, df_a, diff_mask, label=name_a)

    # ── Sheet 2: File B ───────────────────────────────────────────────────
    ws_b = wb.create_sheet(_excel_sheet_name(name_b))
    _write_coloured_sheet(ws_b, df_b, diff_mask, label=name_b)

    # ── Sheet 3: Summary ──────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    _write_summary_sheet(ws_s, stats, diff_mask, name_a, name_b)

    # ── Sheet 4: Mismatch Detail ──────────────────────────────────────────
    ws_m = wb.create_sheet("Mismatch Detail")
    _write_mismatch_detail_sheet(
        ws_m,
        mismatch_list or [],
        name_a, name_b,
        pk_label,
    )

    # ── Sheet 5 (optional): Fuzzy Match Log ───────────────────────────────
    if match_log:
        ws_f = wb.create_sheet("Fuzzy Match Log")
        _write_fuzzy_log_sheet(ws_f, match_log, name_a, name_b)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_html(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    diff_mask: pd.DataFrame,
    stats: dict,
    name_a: str = "File A",
    name_b: str = "File B",
) -> str:
    """Return a self-contained HTML comparison report as a UTF-8 string."""
    compare_cols = [c for c in diff_mask.columns if c != STATUS_COL]

    def _cell_style(row_idx: int, col: str, is_a: bool) -> str:
        if col == STATUS_COL:
            return "background:#fff8dc;color:#555;font-style:italic;"
        if col not in compare_cols:
            return ""
        mismatch = diff_mask.at[row_idx, col] if col in diff_mask.columns else False
        side      = name_a if is_a else name_b
        if mismatch:
            return (
                f"background:#ffc7ce;color:#9c0006;font-weight:600;"
                f"title='Mismatch in {col} (row {row_idx+1}, {side})'"
            )
        return "background:#c6efce;color:#276221;"

    def _build_table(df: pd.DataFrame, title: str, is_a: bool) -> str:
        cols = df.columns.tolist()
        rows_html = []
        for i, row in df.iterrows():
            cells = "".join(
                f'<td style="padding:6px 10px;border:1px solid #ccc;'
                f'{_cell_style(i, col, is_a)}">{_esc(str(row[col]))}</td>'
                for col in cols
            )
            row_num = (
                f'<td style="padding:6px 10px;border:1px solid #ccc;'
                f'background:#f0f0f0;color:#666;font-weight:600;">{i+1}</td>'
            )
            rows_html.append(f"<tr>{row_num}{cells}</tr>")

        header_cells = "".join(
            f'<th style="padding:8px 12px;background:#1f4e79;color:#fff;'
            f'border:1px solid #0d3457;text-align:left;">{_esc(c)}</th>'
            for c in cols
        )
        header = (
            f'<th style="padding:8px 12px;background:#0d3457;color:#fff;'
            f'border:1px solid #0d3457;">#</th>' + header_cells
        )
        return (
            f'<h2 style="color:#1f4e79;margin-top:40px;font-family:sans-serif;">'
            f"{_esc(title)}</h2>"
            f'<div style="overflow-x:auto;">'
            f'<table style="border-collapse:collapse;font-family:monospace;'
            f'font-size:13px;min-width:600px;">'
            f"<thead><tr>{header}</tr></thead>"
            f"<tbody>{''.join(rows_html)}</tbody>"
            f"</table></div>"
        )

    ts      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    table_a = _build_table(df_a, f"📄 {name_a}", is_a=True)
    table_b = _build_table(df_b, f"📄 {name_b}", is_a=False)

    stats_html = (
        '<div style="display:flex;gap:24px;flex-wrap:wrap;margin:24px 0;">'
        + "".join(
            f'<div style="background:#f4f8ff;border:1px solid #c0d4f0;'
            f'border-radius:8px;padding:16px 24px;text-align:center;">'
            f'<div style="font-size:28px;font-weight:700;color:#1f4e79;">{v}</div>'
            f'<div style="color:#555;font-size:12px;margin-top:4px;">{k}</div>'
            f"</div>"
            for k, v in {
                "Total Rows":       stats["total_rows"],
                "Total Cells":      stats["total_cells"],
                "Mismatched Cells": stats["mismatched_cells"],
                "Match Rate":       f"{stats['match_rate_pct']}%",
            }.items()
        )
        + "</div>"
    )

    cols_with_diffs = stats.get("cols_with_diffs", [])
    diff_cols_html  = (
        f'<p style="font-family:sans-serif;color:#9c0006;">'
        f"<strong>Columns with mismatches:</strong> "
        f"{', '.join(_esc(c) for c in cols_with_diffs) if cols_with_diffs else 'None'}</p>"
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"/>
  <meta name="viewport" content="width=device-width,initial-scale=1.0"/>
  <title>Comparison Report — {_esc(name_a)} vs {_esc(name_b)}</title>
  <style>body{{margin:40px;background:#f9fafb;}}h1{{color:#1f4e79;font-family:sans-serif;}}p{{font-family:sans-serif;color:#444;}}</style>
</head>
<body>
  <h1>📊 Comparison Report</h1>
  <p style="color:#888;font-family:sans-serif;">
    <strong>{_esc(name_a)}</strong> vs <strong>{_esc(name_b)}</strong>
    &nbsp;·&nbsp; Generated: {ts}
  </p>
  {stats_html}
  {diff_cols_html}
  <hr style="border:none;border-top:2px solid #ddd;margin:32px 0;"/>
  {table_a}
  {table_b}
  <p style="color:#aaa;font-size:11px;margin-top:40px;">Generated by CSV Validator Tool (Phase 2)</p>
</body>
</html>"""


def to_txt(
    mismatch_list: list[dict],
    stats: dict,
    name_a: str = "File A",
    name_b: str = "File B",
    match_log: list[dict] | None = None,
) -> str:
    """Return a detailed plain-text mismatch summary as a UTF-8 string."""
    lines: list[str] = []
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines += [
        "=" * 72,
        "  COMPARISON REPORT",
        f"  File A : {name_a}",
        f"  File B : {name_b}",
        f"  Generated : {ts}",
        "=" * 72,
        "",
        "SUMMARY",
        "-------",
        f"  Total rows compared : {stats['total_rows']}",
        f"  Total cells compared: {stats['total_cells']}",
        f"  Mismatched cells    : {stats['mismatched_cells']}",
        f"  Match rate          : {stats['match_rate_pct']}%",
        f"  Columns with diffs  : "
        + (", ".join(stats["cols_with_diffs"]) if stats["cols_with_diffs"] else "None"),
        "",
        "=" * 72,
        "MISMATCH DETAILS",
        "=" * 72,
        "",
    ]

    if not mismatch_list:
        lines.append(f"  ✅  No mismatches found — {name_a} and {name_b} are identical "
                     "for the selected columns.")
    else:
        for m in mismatch_list:
            lines += [
                f"  Row      : {m['row']}",
                f"  Column   : {m['column']}",
                f"  {name_a:12s} : {m['value_a']!r}",
                f"  {name_b:12s} : {m['value_b']!r}",
                "-" * 48,
            ]

    # Optional fuzzy match log section
    if match_log:
        fuzzy_rows = [m for m in match_log if m["match_type"] == "fuzzy"]
        unmatched  = [m for m in match_log if "unmatched" in m["match_type"]]
        lines += [
            "",
            "=" * 72,
            "FUZZY MATCH LOG",
            "=" * 72,
            f"  Exact matches  : {sum(1 for m in match_log if m['match_type'] == 'exact')}",
            f"  Fuzzy matches  : {len(fuzzy_rows)}",
            f"  Unmatched rows : {len(unmatched)}",
            "",
        ]
        for m in fuzzy_rows:
            lines += [
                f"  [{name_a}] {m['key_a']!r}",
                f"  [{name_b}] {m['key_b']!r}  — score: {m['score']}",
                "-" * 40,
            ]

    lines += ["", "=" * 72, "END OF REPORT", "=" * 72]
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _write_coloured_sheet(ws, df, diff_mask, label: str) -> None:
    """Write a single coloured data sheet into an openpyxl worksheet."""
    compare_cols = [c for c in diff_mask.columns if c != STATUS_COL]

    # Header row
    ws.cell(row=1, column=1, value="#").fill         = _HEADER_FILL
    ws.cell(row=1, column=1).font                    = _HEADER_FONT
    ws.cell(row=1, column=1).alignment               = Alignment(horizontal="center")
    ws.cell(row=1, column=1).border                  = _THIN_BORDER

    for col_idx, col_name in enumerate(df.columns, start=2):
        cell       = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = _HEADER_FILL
        cell.font  = _HEADER_FONT
        cell.border = _THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        df_row_idx = row_idx - 2  # 0-based index into diff_mask
        status_val = row.get(STATUS_COL, "matched")

        rn_cell       = ws.cell(row=row_idx, column=1, value=df_row_idx + 1)
        rn_cell.fill  = _ROWNR_FILL
        rn_cell.border = _THIN_BORDER
        rn_cell.alignment = Alignment(horizontal="center")

        for col_idx, col_name in enumerate(df.columns, start=2):
            value = row[col_name]
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(wrap_text=True)

            if col_name == STATUS_COL:
                cell.fill = _AMBER_FILL
            elif col_name in compare_cols:
                is_mismatch = (
                    diff_mask.at[df_row_idx, col_name]
                    if df_row_idx in diff_mask.index and col_name in diff_mask.columns
                    else False
                )
                cell.fill = _RED_FILL if is_mismatch else _GREEN_FILL
            elif status_val == "only_in_a" or status_val == "only_in_b":
                cell.fill = _GREY_FILL

    # Auto-size columns
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None), default=8
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 54)

    ws.freeze_panes = "A2"


def _write_summary_sheet(ws, stats, diff_mask, name_a, name_b) -> None:
    """Write the Summary statistics sheet."""
    ws.title = "Summary"

    rows = [
        ("Metric",              "Value"),
        ("File A",              name_a),
        ("File B",              name_b),
        ("Total Rows Compared", stats["total_rows"]),
        ("Total Cells Compared", stats["total_cells"]),
        ("Mismatched Cells",    stats["mismatched_cells"]),
        ("Match Rate (%)",      stats["match_rate_pct"]),
        ("", ""),
        ("Column",              "Mismatch Count"),
    ]

    compare_cols = [c for c in diff_mask.columns if c != STATUS_COL]
    for col in compare_cols:
        rows.append((col, int(diff_mask[col].sum())))

    header_rows = {1, 9}
    for r_idx, (label, value) in enumerate(rows, start=1):
        cl = ws.cell(row=r_idx, column=1, value=label)
        cv = ws.cell(row=r_idx, column=2, value=value)
        if r_idx in header_rows:
            cl.fill = _HEADER_FILL; cv.fill = _HEADER_FILL
            cl.font = _HEADER_FONT; cv.font = _HEADER_FONT
        cl.border = _THIN_BORDER
        cv.border = _THIN_BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


def _write_mismatch_detail_sheet(
    ws,
    mismatch_list: list[dict],
    name_a: str,
    name_b: str,
    pk_label: str,
) -> None:
    """
    Write the Mismatch Detail sheet.

    Column order
    ------------
    Row #  |  Column  |  {pk_label}  |  Value in {name_a}  |  Value in {name_b}

    Formatting
    ----------
    - Header:  navy fill for Row #, Column, and Value columns;
               dark-green fill for the primary-key column.
    - Data:    light-red fill for the two value columns (they are by definition
               mismatches); teal fill for the primary-key column.
    - No-data: a single notice row is written if mismatch_list is empty.
    """
    ws.title = "Mismatch Detail"

    # ── Column header labels ──────────────────────────────────────────────────
    headers = [
        "Row #",
        "Column",
        pk_label,                       # primary/composite key values
        f"Value in {name_a}",
        f"Value in {name_b}",
    ]

    # Header fills — PK column gets a distinct dark-green
    hdr_fills = [
        _HEADER_FILL, _HEADER_FILL,
        _PK_HDR_FILL,
        _HEADER_FILL, _HEADER_FILL,
    ]
    col_widths = [8, 26, 38, 32, 32]   # approximate character widths

    for col_idx, (h, fill, w) in enumerate(zip(headers, hdr_fills, col_widths), start=1):
        cell        = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = fill
        cell.font   = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    if not mismatch_list:
        # Single notice row
        notice = ws.cell(row=2, column=1, value="✅  No mismatches found.")
        notice.font   = Font(color="276221", bold=True, italic=True)
        notice.border = _THIN_BORDER
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        return

    _MISMATCH_VAL_FILL = PatternFill("solid", fgColor="FFCCCC")   # stronger red for values
    _ROWNR_MISMATCH    = PatternFill("solid", fgColor="FCE4D6")   # light salmon for row #

    for r_idx, m in enumerate(mismatch_list, start=2):
        row_num  = m.get("row",        "")
        col_name = m.get("column",     "")
        pk_val   = m.get("primary_key", "N/A")
        val_a    = m.get("value_a",    "")
        val_b    = m.get("value_b",    "")

        row_data = [
            (row_num,  _ROWNR_MISMATCH,    Alignment(horizontal="center")),
            (col_name, _AMBER_FILL,         Alignment()),
            (pk_val,   _TEAL_FILL,          Alignment(wrap_text=True)),
            (val_a,    _MISMATCH_VAL_FILL,  Alignment(wrap_text=True)),
            (val_b,    _MISMATCH_VAL_FILL,  Alignment(wrap_text=True)),
        ]

        for col_idx, (value, fill, align) in enumerate(row_data, start=1):
            cell           = ws.cell(row=r_idx, column=col_idx, value=str(value))
            cell.fill      = fill
            cell.border    = _THIN_BORDER
            cell.alignment = align


def _write_fuzzy_log_sheet(ws, match_log: list[dict], name_a: str, name_b: str) -> None:
    """Write the Fuzzy Match Log sheet."""
    # Colour per match_type
    type_fill = {
        "exact":       _GREEN_FILL,
        "fuzzy":       _BLUE_FILL,
        "unmatched_a": PatternFill("solid", fgColor="FCE4D6"),
        "unmatched_b": PatternFill("solid", fgColor="FCE4D6"),
    }

    headers = [f"Key ({name_a})", f"Key ({name_b})", "Score", "Match Type"]
    for col_idx, h in enumerate(headers, start=1):
        cell      = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    for row_idx, entry in enumerate(match_log, start=2):
        values = [
            entry.get("key_a", ""),
            entry.get("key_b", ""),
            entry.get("score", 0),
            entry.get("match_type", ""),
        ]
        fill = type_fill.get(entry.get("match_type", ""), _GREY_FILL)
        for col_idx, val in enumerate(values, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = fill
            cell.border = _THIN_BORDER

    for i, w in zip("ABCD", [40, 40, 12, 18]):
        ws.column_dimensions[i].width = w
    ws.freeze_panes = "A2"


def _excel_sheet_name(name: str) -> str:
    """Sanitise and truncate a name to fit Excel's 31-char sheet name limit."""
    # Remove characters Excel forbids in sheet names
    for ch in r"\/*?:[]":
        name = name.replace(ch, "_")
    return name[:28] if len(name) > 28 else name


def _esc(text: str) -> str:
    """HTML-escape a string."""
    return (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
    )
